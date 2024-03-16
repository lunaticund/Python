'''
使用此计算器前，需要将 xlsx 文件格式化，第 1 行为标题，第 2 行依次为：
姓名、学号、社会实践队伍名、实践所在省市自治区、社会实践地点、实践时间、队长姓名、立项、团队表彰、个人杰出、校级先进个人（填是否）、校级优秀论文（填是否）、分数，
需保证每名同学的数据仅占 1 行，并将收集表文件放在桌面。
'''

import os
import openpyxl # 此为第三方库，需要安装， Windows 系统下为 pip install openpyxl 命令。


# 以下为计算分数的函数
def calculator (flag, pa, tr, po, slai, slep):
    # flag：年级；pa：立项；tr：团队表彰；po：个人杰出；slai：校级先进个人；slep：校级优秀论文
    score = 0

    # 19、20 级分数计算
    if flag == 19 or flag == 20:

        if pa == '非重点':
            score += 0.5
        elif pa == '院级重点':
            score += 1
        elif pa == '校级一般':
            score += 1.25
        elif pa == '校级重点':
            score += 1.5

        if tr == '无':
            score += 0
        elif tr == '院级优秀':
            score += 0.5
        elif tr == '院级三等奖':
            score += 1
        elif tr == '院级二等奖':
            score += 2
        elif tr == '院级一等奖':
            score += 3
        elif tr == '校十佳':
            score += 6
        
        if po == '无':
            score += 0
        elif po == '个人杰出（团队完成结项）':
            score += 0.5
        elif po == '个人杰出（院级获奖）': # 注意这里是 院级获奖 还是 团队院级获奖
            score += 1
        elif po == '个人杰出（团队获校十佳）':
            score += 3

        if slai == '是':
            score += 2
        elif slai == '否':
            score += 0

        if slep == '是':
            score += 2
        elif slep == '否':
            score += 0

    # 21 级分数计算
    elif flag == 21:
        if l[7].value == '非重点':
            score += 0.5
        elif l[7].value == '院级重点':
            score += 1
        elif l[7].value == '校级一般':
            score += 1.25
        elif l[7].value == '校级重点':
            score += 1.5

        if tr == '无':
            score += 0
        elif tr == '院级优秀':
            score += 0.5
        elif tr == '院级三等奖':
            score += 1
        elif tr == '院级二等奖':
            score += 2
        elif tr == '院级一等奖':
            score += 3
        elif tr == '校十佳':
            score += 6

        if po == '无':
            score += 0
        elif po == '个人杰出（团队完成结项）':
            score += 0.5
        elif po == '个人杰出（院级获奖）': 
            score += 1
        elif po == '个人杰出（团队获校十佳）':
            score += 3

        if slai == '是':
            score += 2
        elif slai == '否':
            score += 0

        if slep == '是':
            score += 2
        elif slep == '否':
            score += 0

    # 22 级分数计算
    elif flag == 22:
        if l[7].value == '非重点':
            score += 0.5
        elif l[7].value == '院级重点':
            score += 1
        elif l[7].value == '校级一般':
            score += 1.25
        elif l[7].value == '校级重点':
            score += 1.5

        if tr == '无':
            score += 0
        elif tr == '院级优秀':
            score += 0.5
        elif tr == '院级三等奖':
            score += 1
        elif tr == '院级二等奖':
            score += 2
        elif tr == '院级一等奖':
            score += 3
        elif tr == '校十佳':
            score += 6

        if po == '无':
            score += 0
        elif po == '个人杰出（团队完成结项）':
            score += 0.5
        elif po == '个人杰出（院级获奖）': 
            score += 1
        elif po == '个人杰出（团队获校十佳）':
            score += 3

        if slai == '是':
            score += 2
        elif slai == '否':
            score += 0

        if slep == '是':
            score += 2
        elif slep == '否':
            score += 0

    return score

# 以下为 main 函数（视作）
path = "C:\\Users\\meiminxuan\\Desktop" # 此处填写电脑桌面路径
os.chdir(path)

workbook = openpyxl.load_workbook('社会实践分数收集表.xlsx') # 此处填写需要读写的 xlsx 文件
sheet = workbook.active
rows = sheet.max_row
columns = sheet.max_column
r = 3

for i in sheet.iter_rows(min_row = 3, max_row = rows, min_col = 1, max_col = columns - 1):
    score = 0
    l = list(i)  
    id = l[1].value
    flag = int(str(id)[1:3])
    score = calculator(flag, l[7].value, l[8].value, l[9].value, l[10].value, l[11].value)
    sheet.cell(r, columns).value = score
    r += 1

workbook.save('社会实践分数收集表.xlsx') # 此处填写需要读写的 xlsx 文件


